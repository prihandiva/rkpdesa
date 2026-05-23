# -*- coding: utf-8 -*-
"""
Script Generator Test Case SIPDES
Mengisi file Excel Test Case dengan 20 modul, 5+ test case per modul
Format mengikuti template yang sudah ada di file Excel asli
"""

import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from copy import copy
import datetime

OUTPUT_PATH = r'd:\laragon\www\RKPDesa\doc gueh\Test Case SIPDES - FINAL.xlsx'
TODAY = datetime.datetime(2025, 7, 31)
REVIEW_DATE = datetime.datetime(2025, 10, 1)

# ─────────────────────────────────────────────
# WARNA & STYLE
# ─────────────────────────────────────────────
BLUE_HEADER   = "1F4E79"   # header baris judul kolom
LIGHT_BLUE    = "BDD7EE"   # baris header info atas
YELLOW_FILL   = "FFD966"   # test scenario #
GREEN_FILL    = "E2EFDA"   # positive
RED_FILL      = "FCE4D6"   # negative
GRAY_FILL     = "D9D9D9"   # sub header
WHITE         = "FFFFFF"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def make_font(bold=False, color="000000", size=9):
    return Font(bold=bold, color=color, size=size)

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

def center_align():
    return Alignment(wrap_text=True, horizontal="center", vertical="center")

# ─────────────────────────────────────────────
# HELPER: tulis baris header info proyek
# ─────────────────────────────────────────────
def write_info_header(ws, module_name):
    info = [
        ("Project Name", "SIPDES"),
        ("Reference Document", "N/A"),
        ("Created By", "Fitria Ramadhani Prihandiva"),
        ("Date of Creation", TODAY),
        ("Date of Review", REVIEW_DATE),
    ]
    for i, (label, val) in enumerate(info, start=1):
        ws.cell(i, 1).value = label
        ws.cell(i, 1).font = make_font(bold=True, size=9)
        ws.cell(i, 2).value = val
        ws.cell(i, 2).font = make_font(size=9)

# ─────────────────────────────────────────────
# HELPER: tulis baris header kolom tabel
# ─────────────────────────────────────────────
COLS = [
    "Page","Test Case ID","Test Type","Test Case Scenario","Test Case",
    "Pre-Conditions","Test Steps","Test Data","Expected Results",
    "Post-Condition","Actual Results","Documentation","Status","PIC","Date",
    "","Bug Description","PIC Dev","Status Dev","Retest Date","Status Retest","PIC Tester"
]

def write_col_header(ws, row=7):
    for c, title in enumerate(COLS, start=1):
        cell = ws.cell(row, c)
        cell.value = title
        cell.fill = make_fill(BLUE_HEADER)
        cell.font = make_font(bold=True, color=WHITE, size=9)
        cell.alignment = center_align()
        cell.border = make_border()

# ─────────────────────────────────────────────
# HELPER: tulis satu baris test case
# ─────────────────────────────────────────────
def write_tc_row(ws, row, page, tc_id, tc_type, scenario, tc_name,
                 pre_cond, steps, data, expected, post_cond):
    fill = make_fill(GREEN_FILL) if tc_type == "Positive" else make_fill(RED_FILL)
    values = [
        page, tc_id, tc_type, scenario, tc_name,
        pre_cond, steps, data, expected, post_cond,
        "", "", "", "", "",  # Actual, Doc, Status, PIC, Date
        "", "", "", "", "", "", ""  # blank | Bug, PIC Dev, Status Dev, Retest, Status Retest, PIC Tester
    ]
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row, c)
        cell.value = val
        cell.border = make_border()
        cell.alignment = wrap_align()
        if c <= 10:
            cell.fill = fill
        if c in (2, 3):
            cell.alignment = center_align()

# ─────────────────────────────────────────────
# SET COLUMN WIDTHS
# ─────────────────────────────────────────────
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

DEFAULT_WIDTHS = [12,14,10,28,32,25,40,25,38,28,20,14,10,10,12,2,30,12,12,12,14,12]

# ─────────────────────────────────────────────
# BUAT WORKBOOK BARU
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # hapus sheet default

# ══════════════════════════════════════════════════════════════
# SHEET 1: TEST SCENARIO
# ══════════════════════════════════════════════════════════════
test_scenarios = [
    (1,  "UC001",       "Periksa Fungsi Login & Logout",
     "1. Login dengan username & password valid\n2. Login dengan password salah\n3. Login dengan field kosong\n4. Akses dashboard tanpa login (proteksi sesi)\n5. Logout dari sistem"),
    (2,  "-",           "Periksa Fungsi Dashboard",
     "1. Tampilan kartu ringkasan data (RPJM, Usulan, RKP)\n2. Navigasi klik kartu menuju modul terkait\n3. Filter data berdasarkan tahun aktif\n4. Dashboard saat data kosong\n5. Cek responsivitas tampilan dashboard"),
    (3,  "-",           "Periksa Fungsi Profil & Pengaturan Akun",
     "1. Edit data profil dengan data valid\n2. Ubah password dengan password baru valid\n3. Ubah password dengan konfirmasi tidak cocok\n4. Upload foto profil format valid (JPG/PNG)\n5. Upload foto profil format tidak valid"),
    (4,  "-",           "Periksa Fungsi Notifikasi",
     "1. Lihat daftar notifikasi masuk\n2. Tandai satu notifikasi sebagai sudah dibaca\n3. Tandai semua notifikasi sebagai sudah dibaca\n4. Tampilan notifikasi saat tidak ada notifikasi baru\n5. Klik notifikasi menuju halaman yang relevan"),
    (5,  "UC008-UC011", "Periksa Fungsi RPJM Desa (CRUD + Cetak)",
     "1. Tambah data RPJM Desa dengan data lengkap\n2. Tambah RPJM tanpa mengisi field wajib\n3. Edit/Ubah data RPJM Desa\n4. Hapus data RPJM Desa (konfirmasi Ya)\n5. Cetak/Export data RPJM Desa ke PDF"),
    (6,  "UC001-UC004", "Periksa Fungsi Usulan (CRUD + Checklist RKP)",
     "1. Tambah data usulan dengan data lengkap\n2. Tambah usulan tanpa mengisi field wajib\n3. Edit/Ubah data usulan\n4. Hapus data usulan (konfirmasi Ya)\n5. Checklist usulan layak masuk ke RKP Desa"),
    (7,  "UC016-UC019", "Periksa Fungsi RKP Desa (CRUD + Cetak + Approval)",
     "1. Input/Tambah data RKP Desa dari hasil verifikasi\n2. Edit/Ubah data RKP Desa\n3. Hapus data RKP Desa\n4. Cetak/Export dokumen RKP Desa ke PDF\n5. Approval/Persetujuan RKP Desa oleh BPD"),
    (8,  "UC005-UC007", "Periksa Fungsi Berita Acara (Lihat + Unggah + Cetak)",
     "1. Lihat daftar dan detail Berita Acara\n2. Unggah dokumen Berita Acara format PDF\n3. Unggah dokumen format tidak valid (bukan PDF)\n4. Cetak Berita Acara\n5. Download dokumen Berita Acara"),
    (9,  "UC013-UC015", "Periksa Fungsi Verifikasi Usulan",
     "1. Input hasil verifikasi status Layak\n2. Input hasil verifikasi status Tidak Layak dengan catatan\n3. Input hasil verifikasi Tidak Layak tanpa catatan (field kosong)\n4. Edit/Ubah hasil verifikasi\n5. Lihat daftar hasil verifikasi"),
    (10, "UC002",       "Periksa Fungsi Manajemen Tahun (CRUD)",
     "1. Tambah tahun perencanaan baru yang valid\n2. Tambah tahun dengan data duplikat\n3. Edit/Ubah status aktif/nonaktif tahun\n4. Hapus data tahun yang tidak terpakai\n5. Tambah tahun tanpa mengisi form"),
    (11, "UC003",       "Periksa Fungsi Manajemen Pengguna (CRUD)",
     "1. Tambah pengguna baru dengan data lengkap & valid\n2. Tambah pengguna dengan username yang sudah ada\n3. Edit/Ubah data dan role pengguna\n4. Reset password pengguna\n5. Hapus data pengguna"),
    (12, "-",           "Periksa Fungsi Master Bidang (CRUD)",
     "1. Tambah data Bidang baru\n2. Tambah data Bidang dengan nama duplikat\n3. Edit/Ubah nama Bidang\n4. Hapus Bidang yang tidak digunakan\n5. Tambah Bidang tanpa mengisi nama"),
    (13, "-",           "Periksa Fungsi Master Sumber Dana (CRUD)",
     "1. Tambah data Sumber Dana baru\n2. Tambah Sumber Dana dengan nama duplikat\n3. Edit/Ubah nama Sumber Dana\n4. Hapus Sumber Dana yang tidak digunakan\n5. Tambah Sumber Dana tanpa mengisi nama"),
    (14, "-",           "Periksa Fungsi Master Pola Pelaksanaan (CRUD)",
     "1. Tambah data Pola Pelaksanaan baru\n2. Tambah Pola Pelaksanaan dengan nama duplikat\n3. Edit/Ubah nama Pola Pelaksanaan\n4. Hapus Pola Pelaksanaan yang tidak digunakan\n5. Tambah Pola Pelaksanaan tanpa mengisi nama"),
    (15, "-",           "Periksa Fungsi Master Dusun (CRUD)",
     "1. Tambah data Dusun baru dengan nama valid\n2. Tambah Dusun dengan nama duplikat\n3. Edit/Ubah nama Dusun\n4. Hapus Dusun yang tidak memiliki data terkait\n5. Tambah Dusun tanpa mengisi nama"),
    (16, "-",           "Periksa Fungsi Master RW (CRUD)",
     "1. Tambah data RW baru (nomor & dusun valid)\n2. Tambah RW dengan nomor duplikat pada dusun yang sama\n3. Edit/Ubah nomor atau dusun induk RW\n4. Hapus data RW yang tidak memiliki RT terkait\n5. Tambah RW tanpa mengisi field wajib"),
    (17, "-",           "Periksa Fungsi Master RT (CRUD)",
     "1. Tambah data RT baru (nomor & RW valid)\n2. Tambah RT dengan nomor duplikat pada RW yang sama\n3. Edit/Ubah nomor atau RW induk RT\n4. Hapus data RT yang tidak memiliki warga/usulan terkait\n5. Tambah RT tanpa mengisi field wajib"),
    (18, "-",           "Periksa Fungsi Master Pegawai (CRUD)",
     "1. Tambah data Pegawai baru dengan data lengkap\n2. Tambah Pegawai dengan NIP yang sudah ada\n3. Edit/Ubah data Pegawai\n4. Hapus data Pegawai\n5. Tambah Pegawai tanpa mengisi field wajib"),
    (19, "UC004",       "Periksa Fungsi Monitoring Sistem",
     "1. Buka dan tampilkan tabel log aktivitas pengguna\n2. Filter log berdasarkan tanggal tertentu\n3. Filter log berdasarkan nama/role pengguna\n4. Lihat detail aktivitas (klik baris log)\n5. Coba filter dengan format tanggal tidak valid"),
    (20, "-",           "Periksa Fungsi Pemulihan Data (Soft Delete Recovery)",
     "1. Lihat daftar data yang telah dihapus (soft delete)\n2. Pulihkan satu data yang telah dihapus\n3. Hapus permanen satu data dari daftar pemulihan\n4. Pulihkan semua data sekaligus (bulk restore)\n5. Cari data terhapus menggunakan fitur pencarian"),
]

ws_ts = wb.create_sheet("Test Scenario")
# Header info
ts_info = [("Project Name","SIPDES"),("Reference Document","N/A"),
           ("Created By","Fitria Ramadhani Prihandiva"),
           ("Date of Creation",TODAY),("Date of Review",REVIEW_DATE)]
for i,(l,v) in enumerate(ts_info,1):
    ws_ts.cell(i,1).value=l; ws_ts.cell(i,1).font=make_font(bold=True)
    ws_ts.cell(i,2).value=v

# Header kolom
hdr_titles = ["Test Scenario #","Requirement ID (Kode UC)","Test Scenario Description","Test Cases"]
for c,t in enumerate(hdr_titles,1):
    cell=ws_ts.cell(7,c); cell.value=t
    cell.fill=make_fill(BLUE_HEADER); cell.font=make_font(bold=True,color=WHITE)
    cell.alignment=center_align(); cell.border=make_border()

# Isi data
for r,scen in enumerate(test_scenarios, start=8):
    num,req_id,desc,cases = scen
    ws_ts.cell(r,1).value=num; ws_ts.cell(r,1).fill=make_fill(YELLOW_FILL)
    ws_ts.cell(r,1).alignment=center_align(); ws_ts.cell(r,1).border=make_border()
    ws_ts.cell(r,1).font=make_font(bold=True)
    ws_ts.cell(r,2).value=req_id; ws_ts.cell(r,2).alignment=center_align()
    ws_ts.cell(r,2).border=make_border()
    ws_ts.cell(r,3).value=desc; ws_ts.cell(r,3).alignment=wrap_align()
    ws_ts.cell(r,3).border=make_border(); ws_ts.cell(r,3).font=make_font(bold=True)
    ws_ts.cell(r,4).value=cases; ws_ts.cell(r,4).alignment=wrap_align()
    ws_ts.cell(r,4).border=make_border()
    ws_ts.row_dimensions[r].height=90

ws_ts.column_dimensions['A'].width=8
ws_ts.column_dimensions['B'].width=18
ws_ts.column_dimensions['C'].width=38
ws_ts.column_dimensions['D'].width=70
ws_ts.row_dimensions[7].height=20
print("Sheet 'Test Scenario' selesai.")
